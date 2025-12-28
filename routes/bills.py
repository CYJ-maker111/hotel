from flask import Blueprint, request, jsonify, send_file
import sqlite3
import os
import csv
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from app_context import system

# 尝试导入openpyxl，如果没有安装则在导出时提示
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# 创建账单管理蓝图
bills_bp = Blueprint('bills', __name__, url_prefix='/api/bills')

@bills_bp.route('/<int:room_id>/detail', methods=['GET'])
def get_room_bill_detail(room_id):
    """
    获取指定房间的空调使用详单
    字段：房间号、请求时间、服务开始时间、服务结束时间、服务时长（秒）、风速、当前费用、累积费用
    使用room.cost作为总费用，确保与管理员界面一致
    """
    # 连接数据库
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'hotel_ac.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        # 使用room.cost作为总费用
        room = system.scheduler.rooms.get(room_id)
        if not room:
            return jsonify({
                "status": "error",
                "message": f"房间{room_id}不存在"
            }), 404
        
        total_cost = room.cost
        
        # 查询房间的详单记录（包含新字段）
        cursor.execute("""
            SELECT room_id, request_time, start_time, end_time, service_duration, 
                   fan_speed, cost, accumulated_cost, mode, target_temp, operation_type
            FROM detail_records 
            WHERE room_id = ? 
            ORDER BY start_time ASC
        """, (room_id,))
        
        records = cursor.fetchall()
        
        # 转换为JSON格式，费用完全基于room.cost，不使用数据库中的费用数据
        result = []
        if records:
            # 完全基于room.cost计算费用，不使用数据库中的费用数据
            # 将room.cost按服务时长比例分配给各条记录
            
            # 计算总服务时长（用于按比例分配费用）
            total_duration = sum(int(record[4] or 0) for record in records)
            
            # 按时间顺序计算累积费用（从最早到最晚）
            accumulated = 0.0
            for i, record in enumerate(records):  # 已经按时间正序排列
                # 如果是最后一条记录，确保累积费用等于room.cost
                if i == len(records) - 1:
                    # 最后一条记录的累积费用必须等于total_cost
                    if i > 0:
                        # 前面所有记录的累积费用
                        prev_accumulated = accumulated
                        # 最后一条记录的费用 = total_cost - 前面所有记录的累积费用
                        adjusted_cost = round(total_cost - prev_accumulated, 2)
                        accumulated = round(total_cost, 2)
                    else:
                        # 只有一条记录
                        adjusted_cost = round(total_cost, 2)
                        accumulated = round(total_cost, 2)
                else:
                    # 非最后一条记录，按服务时长比例分配费用
                    record_duration = int(record[4] or 0)
                    if total_duration > 0:
                        # 按服务时长比例分配
                        adjusted_cost = round(total_cost * (record_duration / total_duration), 2)
                    else:
                        # 如果总时长为0，按记录数量平均分配
                        adjusted_cost = round(total_cost / len(records), 2)
                    accumulated += adjusted_cost
                    accumulated = round(accumulated, 2)
                
                result.append({
                    "room_id": record[0],           # 房间号
                    "request_time": record[1] or record[2] or '',  # 请求时间（若无则使用开始时间）
                    "start_time": record[2] or '',        # 服务开始时间
                    "end_time": record[3] or '',          # 服务结束时间
                    "service_duration": int(record[4] or 0),  # 服务时长（秒）
                    "fan_speed": record[5] or '',         # 风速
                    "cost": round(adjusted_cost, 2),    # 当前费用（基于room.cost计算，不使用数据库费用）
                    "accumulated_cost": round(accumulated, 2),  # 累积费用（基于room.cost）
                    # 额外字段
                    "mode": record[8] or '',
                    "target_temp": float(record[9] or 0),
                    "operation_type": record[10] or ''
                })
            
            # 结果已经是按时间正序，但前端可能需要倒序显示，所以反转
            result.reverse()
        else:
            # 如果没有记录，返回空列表
            pass
        
        return jsonify({
            "room_id": room_id,
            "records": result,
            "total_records": len(result)
        })
    except Exception as e:
        return jsonify({"status": "error", "message": f"获取详单失败：{str(e)}"}), 500
    finally:
        conn.close()

@bills_bp.route('/<int:room_id>/detail/export_csv', methods=['GET'])
def export_detail_csv(room_id):
    """
    导出指定房间的空调使用详单为CSV文件
    """
    # 连接数据库
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'hotel_ac.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        # 获取房间信息（用于文件名）
        cursor.execute("""
            SELECT guest_name, checkin_time
            FROM checkin_records 
            WHERE room_id = ? AND status = 'CHECKED_IN'
            ORDER BY checkin_time DESC
            LIMIT 1
        """, (room_id,))
        
        checkin_info = cursor.fetchone()
        guest_name = checkin_info[0] if checkin_info else ""
        
        # 查询房间的详单记录
        cursor.execute("""
            SELECT room_id, request_time, start_time, end_time, service_duration, 
                   fan_speed, cost, accumulated_cost, mode, target_temp, operation_type
            FROM detail_records 
            WHERE room_id = ? 
            ORDER BY start_time ASC
        """, (room_id,))
        
        records = cursor.fetchall()
        
        # 创建CSV内容
        output = StringIO()
        # 使用quoting=csv.QUOTE_ALL确保所有字段都用引号包裹，避免Excel解析错误
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)
        
        # 写入BOM（用于Excel正确显示中文）
        output.write('\ufeff')
        
        # 写入表头
        headers = ["房间号", "请求时间", "开始时间", "结束时间", "服务时长(秒)", 
                   "风速", "费用(元)", "累积费用(元)", "模式", "目标温度(℃)", "操作类型"]
        writer.writerow(headers)
        
        # 写入数据
        for record in records:
            # 格式化日期时间，确保格式正确
            request_time = record[1] or record[2] or ''
            start_time = record[2] or ''
            end_time = record[3] or ''
            
            # 确保日期时间格式为 YYYY-MM-DD HH:MM:SS（Excel能识别的格式）
            # 如果已经是正确格式，直接使用；否则尝试转换
            try:
                if request_time and len(request_time) > 10:
                    # 如果包含时间部分，确保格式正确
                    if 'T' in request_time:
                        request_time = request_time.replace('T', ' ')
                    if '.' in request_time:
                        request_time = request_time.split('.')[0]
            except:
                pass
            
            try:
                if start_time and len(start_time) > 10:
                    if 'T' in start_time:
                        start_time = start_time.replace('T', ' ')
                    if '.' in start_time:
                        start_time = start_time.split('.')[0]
            except:
                pass
            
            try:
                if end_time and len(end_time) > 10:
                    if 'T' in end_time:
                        end_time = end_time.replace('T', ' ')
                    if '.' in end_time:
                        end_time = end_time.split('.')[0]
            except:
                pass
            
            writer.writerow([
                record[0],  # room_id
                request_time,  # request_time
                start_time,  # start_time
                end_time,  # end_time
                record[4] or 0,  # service_duration
                record[5] or '',  # fan_speed
                round(record[6], 2) if record[6] else 0,  # cost
                round(record[7], 2) if record[7] else 0,  # accumulated_cost
                record[8] or '',  # mode
                record[9] or '',  # target_temp
                record[10] or ''  # operation_type
            ])
        
        # 准备文件内容
        output.seek(0)
        csv_content = output.getvalue()
        output.close()
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"房间{room_id}_空调使用详单_{timestamp}.csv"
        if guest_name:
            filename = f"房间{room_id}_{guest_name}_空调使用详单_{timestamp}.csv"
        
        # 返回CSV文件
        return send_file(
            BytesIO(csv_content.encode('utf-8-sig')),  # utf-8-sig包含BOM，Excel可以正确打开
            mimetype='text/csv; charset=utf-8',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({"status": "error", "message": f"导出CSV失败：{str(e)}"}), 500
    finally:
        conn.close()

@bills_bp.route('/<int:room_id>/summary', methods=['GET'])
def get_room_bill_summary(room_id):
    """
    获取指定房间的账单汇总信息
    使用room.cost（内存中的费用）而不是数据库中的费用
    """
    # 连接数据库
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'hotel_ac.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        # 使用room.cost（内存中的费用）
        room = system.scheduler.rooms.get(room_id)
        total_cost = room.cost if room else 0.0
        
        # 查询总时长（从数据库）
        cursor.execute("""
            SELECT SUM(julianday(end_time) - julianday(start_time)) * 86400
            FROM detail_records 
            WHERE room_id = ?
        """, (room_id,))
        
        result = cursor.fetchone()
        total_duration = result[0] or 0  # 总时长（秒）
        
        # 获取当前入住信息
        cursor.execute("""
            SELECT guest_name, checkin_time, checkout_time 
            FROM checkin_records 
            WHERE room_id = ? AND status = 'CHECKED_IN'
        """, (room_id,))
        
        checkin_info = cursor.fetchone()
        guest_info = {
            "guest_name": checkin_info[0] if checkin_info else "",
            "checkin_time": checkin_info[1] if checkin_info else "",
            "checkout_time": checkin_info[2] if checkin_info else ""
        }
        
        return jsonify({
            "room_id": room_id,
            "total_cost": round(total_cost, 2),
            "total_duration": round(total_duration, 2),
            "guest_info": guest_info
        })
    except Exception as e:
        return jsonify({"status": "error", "message": f"获取账单汇总失败：{str(e)}"}), 500
    finally:
        conn.close()

@bills_bp.route('/<int:room_id>/ac_bill', methods=['GET'])
def get_ac_bill(room_id):
    """
    获取空调账单（仅空调费用）
    字段：房间号、客人姓名、入住时间、结束时间（入住时间+天数）、空调总费用
    不计算关机次数，只显示空调使用费用
    """
    # 连接数据库
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'hotel_ac.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        # 获取入住信息（优先获取当前入住的，如果没有则获取最近的一条记录）
        cursor.execute("""
            SELECT guest_name, checkin_time, power_off_count
            FROM checkin_records 
            WHERE room_id = ?
            ORDER BY 
                CASE WHEN status = 'CHECKED_IN' THEN 0 ELSE 1 END,
                checkin_time DESC
            LIMIT 1
        """, (room_id,))
        
        checkin_info = cursor.fetchone()
        
        if not checkin_info:
            return jsonify({"status": "error", "message": "该房间没有入住记录"}), 404
        
        guest_name, checkin_time, power_off_count = checkin_info
        power_off_count = power_off_count or 0
        days = max(power_off_count, 1)
        
        # 计算结束时间 = 入住时间 + 天数
        try:
            checkin_dt = datetime.strptime(checkin_time, "%Y-%m-%d %H:%M:%S")
            end_time = checkin_dt + timedelta(days=days)
            end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        except:
            end_time_str = ""
        
        # 使用room.cost（内存中的费用）作为空调总费用
        room = system.scheduler.rooms.get(room_id)
        ac_total_cost = room.cost if room else 0.0
        
        return jsonify({
            "bill_type": "空调账单",
            "room_id": room_id,
            "guest_name": guest_name,
            "checkin_time": checkin_time,
            "end_time": end_time_str,
            "days": days,
            "ac_total_cost": round(ac_total_cost, 2),
            "total_cost": round(ac_total_cost, 2)
        })
    except Exception as e:
        return jsonify({"status": "error", "message": f"获取空调账单失败：{str(e)}"}), 500
    finally:
        conn.close()

@bills_bp.route('/<int:room_id>/accommodation_bill', methods=['GET'])
def get_accommodation_bill(room_id):
    """
    获取系统住宿账单（住宿费+空调费）
    字段：房间号、入住时间、结束时间、空调总费用、住宿总费用、总计
    房间费用：R1=100元/天，R2=125元/天，R3=150元/天，R4=200元/天，R5=100元/天
    每次关机视为过了一天
    """
    # 连接数据库
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'hotel_ac.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        # 定义房间日租金
        room_rates = {1: 100, 2: 125, 3: 150, 4: 200, 5: 100}
        daily_rate = room_rates.get(room_id, 100)
        
        # 获取入住信息（优先获取当前入住的，如果没有则获取最近的一条记录）
        cursor.execute("""
            SELECT guest_name, checkin_time, checkout_time, power_off_count
            FROM checkin_records 
            WHERE room_id = ?
            ORDER BY 
                CASE WHEN status = 'CHECKED_IN' THEN 0 ELSE 1 END,
                checkin_time DESC
            LIMIT 1
        """, (room_id,))
        
        checkin_info = cursor.fetchone()
        
        if not checkin_info:
            return jsonify({"status": "error", "message": "该房间没有入住记录"}), 404
        
        guest_name, checkin_time, checkout_time, power_off_count = checkin_info
        power_off_count = power_off_count or 0
        days = max(power_off_count, 1)
        
        # 计算结束时间 = 入住时间 + 天数
        try:
            checkin_dt = datetime.strptime(checkin_time, "%Y-%m-%d %H:%M:%S")
            end_time = checkin_dt + timedelta(days=days)
            end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        except:
            end_time_str = checkout_time or ""
        
        # 计算住宿费用：按关机次数计算天数
        accommodation_cost = daily_rate * days
        
        # 使用room.cost（内存中的费用）作为空调总费用
        room = system.scheduler.rooms.get(room_id)
        ac_total_cost = room.cost if room else 0.0
        
        # 计算总费用
        total_cost = accommodation_cost + ac_total_cost
        
        return jsonify({
            "bill_type": "系统住宿账单",
            "room_id": room_id,
            "guest_name": guest_name,
            "checkin_time": checkin_time,
            "end_time": end_time_str,
            "days": days,
            "daily_rate": daily_rate,
            "accommodation_cost": round(accommodation_cost, 2),
            "ac_total_cost": round(ac_total_cost, 2),
            "total_cost": round(total_cost, 2)
        })
    except Exception as e:
        return jsonify({"status": "error", "message": f"获取综合账单失败：{str(e)}"}), 500
    finally:
        conn.close()

@bills_bp.route('/<int:room_id>/reset', methods=['POST'])
def reset_room_cost(room_id):
    """
    重置指定房间的累计费用（清空该房间的所有详单记录）
    用于换客人时清零累计费用
    """
    # 连接数据库
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'hotel_ac.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        # 删除该房间的所有详单记录
        cursor.execute("DELETE FROM detail_records WHERE room_id = ?", (room_id,))
        deleted_count = cursor.rowcount
        
        # 重置房间的当前费用
        room = system.scheduler.rooms.get(room_id)
        if room:
            room.cost = 0.0
        
        # 清除当前记录ID
        if room_id in system.scheduler.current_record_ids:
            del system.scheduler.current_record_ids[room_id]
        
        conn.commit()
        
        return jsonify({
            "status": "success",
            "message": f"房间{room_id}的累计费用已重置",
            "deleted_records": deleted_count
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": f"重置失败：{str(e)}"}), 500
    finally:
        conn.close()

@bills_bp.route('/<int:room_id>/export', methods=['GET'])
def export_room_bill(room_id):
    """
    导出指定房间的账单为Excel文件
    参数：bill_type (query parameter) - 可选值：ac（空调账单）或 accommodation（系统住宿账单）
    """
    if not EXCEL_AVAILABLE:
        return jsonify({"status": "error", "message": "服务器未安装openpyxl库，无法导出Excel"}), 500
    
    bill_type = request.args.get('bill_type', 'accommodation')
    
    try:
        # 连接数据库
        db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'hotel_ac.db')
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # 定义房间日租金
        room_rates = {1: 100, 2: 125, 3: 150, 4: 200, 5: 100}
        daily_rate = room_rates.get(room_id, 100)
        
        # 获取入住信息（优先获取当前入住的，如果没有则获取最近的一条记录）
        cursor.execute("""
            SELECT guest_name, checkin_time, checkout_time, power_off_count
            FROM checkin_records 
            WHERE room_id = ?
            ORDER BY 
                CASE WHEN status = 'CHECKED_IN' THEN 0 ELSE 1 END,
                checkin_time DESC
            LIMIT 1
        """, (room_id,))
        
        checkin_info = cursor.fetchone()
        
        if not checkin_info:
            conn.close()
            return jsonify({"status": "error", "message": "该房间没有入住记录"}), 404
        
        guest_name, checkin_time, checkout_time, power_off_count = checkin_info
        power_off_count = power_off_count or 0
        days = max(power_off_count, 1)
        
        # 计算结束时间 = 入住时间 + 天数
        try:
            checkin_dt = datetime.strptime(checkin_time, "%Y-%m-%d %H:%M:%S")
            end_time = checkin_dt + timedelta(days=days)
            end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        except:
            end_time_str = checkout_time or ""
        
        # 使用room.cost（内存中的费用）作为空调总费用
        room = system.scheduler.rooms.get(room_id)
        ac_total_cost = room.cost if room else 0.0
        
        # 构建账单数据
        if bill_type == 'ac':
            bill_data = {
                "bill_type": "空调账单",
                "room_id": room_id,
                "guest_name": guest_name,
                "checkin_time": checkin_time,
                "end_time": end_time_str,
                "days": days,
                "ac_total_cost": round(ac_total_cost, 2),
                "total_cost": round(ac_total_cost, 2)
            }
        else:
            accommodation_cost = daily_rate * days
            bill_data = {
                "bill_type": "系统住宿账单",
                "room_id": room_id,
                "guest_name": guest_name,
                "checkin_time": checkin_time,
                "end_time": end_time_str,
                "days": days,
                "daily_rate": daily_rate,
                "accommodation_cost": round(accommodation_cost, 2),
                "ac_total_cost": round(ac_total_cost, 2),
                "total_cost": round(accommodation_cost + ac_total_cost, 2)
            }
        
        # 获取详单记录
        cursor.execute("""
            SELECT request_time, start_time, end_time, service_duration, 
                   fan_speed, cost, accumulated_cost, mode, target_temp, operation_type
            FROM detail_records 
            WHERE room_id = ? 
            ORDER BY start_time ASC
        """, (room_id,))
        
        details = cursor.fetchall()
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"房间{room_id}账单"
        
        # 设置样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # 写入账单摘要
        ws.append([f"{bill_data.get('bill_type', '账单')} - 房间{room_id}"])
        ws.merge_cells('A1:G1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.append([])
        ws.append(["客人姓名", bill_data.get('guest_name', '')])
        ws.append(["入住时间", bill_data.get('checkin_time', '')])
        ws.append(["结束时间", bill_data.get('end_time', '')])
        ws.append(["住宿天数", bill_data.get('days', 0)])
        
        if bill_type == 'accommodation':
            ws.append(["房间日租金", f"¥{bill_data.get('daily_rate', 0)}"])
            ws.append(["住宿费用", f"¥{bill_data.get('accommodation_cost', 0)}"])
        
        ws.append(["空调费用", f"¥{bill_data.get('ac_total_cost', 0)}"])
        ws.append(["总计", f"¥{bill_data.get('total_cost', 0)}"])
        
        # 空一行
        ws.append([])
        
        # 写入详单标题
        ws.append(["空调使用详单"])
        ws.merge_cells(f'A{ws.max_row}:J{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='center')
        
        # 写入详单表头
        headers = ["请求时间", "开始时间", "结束时间", "时长(秒)", "风速", "费用", "累积费用", "模式", "目标温度", "操作类型"]
        header_row = ws.max_row + 1
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 写入详单数据
        for record in details:
            ws.append([
                record[0] or record[1],  # request_time
                record[1],  # start_time
                record[2] or '-',  # end_time
                record[3] or 0,  # service_duration
                record[4],  # fan_speed
                round(record[5], 2),  # cost
                round(record[6], 2),  # accumulated_cost
                record[7],  # mode
                record[8],  # target_temp
                record[9]  # operation_type
            ])
        
        # 调整列宽
        for idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(idx)
            for cell in col:
                try:
                    # 跳过合并单元格
                    if hasattr(cell, 'value'):
                        cell_value = str(cell.value) if cell.value is not None else ''
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 50)  # 最小10，最大50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        conn.close()
        
        # 生成文件名
        bill_type_cn = "空调账单" if bill_type == 'ac' else "系统住宿账单"
        filename = f"房间{room_id}_{bill_type_cn}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({"status": "error", "message": f"导出账单失败：{str(e)}"}), 500